"""Tests for upsert_notebooklm_index workbook upsert tool."""

import importlib.util
from pathlib import Path
from types import ModuleType

import openpyxl
import pytest

REPO_ROOT = Path(__file__).resolve().parents[3]
TOOLS_DIR = REPO_ROOT / ".agents" / "tools" / "library"


def load_upsert_module() -> ModuleType:
    module_path = TOOLS_DIR / "upsert_notebooklm_index.py"
    spec = importlib.util.spec_from_file_location(
        "upsert_notebooklm_index", module_path
    )
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Could not load: {module_path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


NB_HEADERS = [
    "notebook_id",
    "title",
    "url",
    "source_count",
    "summary",
    "suggested_topics",
    "last_updated",
    "status",
]
SRC_HEADERS = [
    "notebook_id",
    "notebook_title",
    "source_id",
    "source_title",
    "source_summary",
    "source_keywords",
]


@pytest.fixture
def tmp_index(tmp_path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws_nb = wb.active
    ws_nb.title = "notebooks"
    ws_nb.append(NB_HEADERS)
    ws_src = wb.create_sheet("sources")
    ws_src.append(SRC_HEADERS)
    index_path = tmp_path / "index-notebooklm.xlsx"
    wb.save(index_path)
    return index_path


@pytest.fixture
def upsert_mod(tmp_index: Path, monkeypatch: pytest.MonkeyPatch) -> ModuleType:
    mod = load_upsert_module()
    monkeypatch.setattr(mod, "INDEX_PATH", tmp_index)
    return mod


def _make_notebook(notebook_id: str = "nb-1") -> dict:
    return {
        "notebook_id": notebook_id,
        "title": "Test Notebook",
        "url": "https://notebooklm.google.com/notebook/nb-1",
        "source_count": 2,
        "summary": "A summary",
        "suggested_topics": "topic1; topic2",
    }


def _make_source(source_id: str = "src-1") -> dict:
    return {
        "source_id": source_id,
        "source_title": "Source Title",
        "source_summary": "Source summary",
        "source_keywords": "kw1; kw2",
    }


def test_upsert_inserts_new_notebook_row(
    upsert_mod: ModuleType, tmp_index: Path
) -> None:
    # Arrange
    notebook = _make_notebook()
    sources = [_make_source()]

    # Act
    upsert_mod.upsert(notebook, sources)

    # Assert
    wb = openpyxl.load_workbook(tmp_index)
    assert wb["notebooks"].max_row == 2  # header + 1 data row
    assert wb["sources"].max_row == 2
    assert wb["notebooks"].cell(row=2, column=1).value == "nb-1"


def test_upsert_overwrites_existing_notebook_row(
    upsert_mod: ModuleType, tmp_index: Path
) -> None:
    # Arrange
    notebook = _make_notebook()
    sources = [_make_source()]
    upsert_mod.upsert(notebook, sources)

    # Act — upsert again with updated summary
    notebook["summary"] = "Updated summary"
    upsert_mod.upsert(notebook, sources)

    # Assert — still only one data row, with new summary
    wb = openpyxl.load_workbook(tmp_index)
    assert wb["notebooks"].max_row == 2
    assert wb["notebooks"].cell(row=2, column=5).value == "Updated summary"


def test_upsert_replaces_old_sources_with_new(
    upsert_mod: ModuleType, tmp_index: Path
) -> None:
    # Arrange
    notebook = _make_notebook()
    upsert_mod.upsert(notebook, [_make_source("src-1"), _make_source("src-2")])

    # Act — re-upsert with only one source
    upsert_mod.upsert(notebook, [_make_source("src-new")])

    # Assert
    wb = openpyxl.load_workbook(tmp_index)
    assert wb["sources"].max_row == 2  # header + 1 source
    assert wb["sources"].cell(row=2, column=3).value == "src-new"


def test_upsert_tolerates_missing_optional_fields(
    upsert_mod: ModuleType, tmp_index: Path
) -> None:
    # Arrange — omit summary, suggested_topics, source_summary, source_keywords
    notebook = {
        "notebook_id": "nb-2",
        "title": "Sparse Notebook",
        "url": "https://notebooklm.google.com/notebook/nb-2",
        "source_count": 1,
    }
    source = {"source_id": "src-x", "source_title": "Sparse Source"}

    # Act — must not raise KeyError
    upsert_mod.upsert(notebook, [source])

    # Assert — row was written without raising KeyError; openpyxl returns None for empty cells
    wb = openpyxl.load_workbook(tmp_index)
    assert wb["notebooks"].max_row == 2
    assert wb["sources"].max_row == 2
    assert wb["notebooks"].cell(row=2, column=1).value == "nb-2"
