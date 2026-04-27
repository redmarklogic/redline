"""Tests for library workbook maintenance helpers."""

import importlib.util
from pathlib import Path
from types import ModuleType

import openpyxl
import pytest

REPO_ROOT = Path(__file__).resolve().parents[3]
TOOLS_DIR = REPO_ROOT / ".agents" / "tools" / "library"


def load_tool_module(module_name: str) -> ModuleType:
    module_path = TOOLS_DIR / f"{module_name}.py"
    spec = importlib.util.spec_from_file_location(module_name, module_path)
    if spec is None or spec.loader is None:
        message = f"Could not load tool module: {module_name}"
        raise RuntimeError(message)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


@pytest.fixture
def workbook_utils() -> ModuleType:
    return load_tool_module("workbook_utils")


@pytest.fixture
def enrich_safe_metadata() -> ModuleType:
    return load_tool_module("enrich_safe_metadata")


@pytest.fixture
def verify_index() -> ModuleType:
    return load_tool_module("verify_index")


@pytest.fixture
def export_needs_review() -> ModuleType:
    return load_tool_module("export_needs_review")


@pytest.fixture
def export_review_pack() -> ModuleType:
    return load_tool_module("export_review_pack")


def test_workbook_lock_fails_when_another_writer_is_active(
    tmp_path: Path, workbook_utils: ModuleType
) -> None:
    index_path = tmp_path / "library-index.xlsx"

    with (
        workbook_utils.WorkbookLock(index_path),
        pytest.raises(RuntimeError, match="Another library index writer is active"),
        workbook_utils.WorkbookLock(index_path),
    ):
        pass

    assert not (tmp_path / "library-index.xlsx.lock").exists()


def test_save_workbook_atomically_replaces_target(
    tmp_path: Path, workbook_utils: ModuleType
) -> None:
    target_path = tmp_path / "library-index.xlsx"
    target_path.write_text("old", encoding="utf-8")

    class FakeWorkbook:
        def save(self, path: Path) -> None:
            path.write_text("new", encoding="utf-8")

    workbook_utils.save_workbook_atomically(FakeWorkbook(), target_path)

    assert target_path.read_text(encoding="utf-8") == "new"
    assert list(tmp_path.glob("*.tmp.xlsx")) == []


def test_append_index_row_writes_master_and_domain_rows(
    workbook_utils: ModuleType,
) -> None:
    # Arrange
    workbook = openpyxl.Workbook()
    master = workbook.active
    master.title = "Master"
    master.append(workbook_utils.STANDARD_HEADERS)
    domain = workbook.create_sheet("Standards")
    domain.append(
        workbook_utils.STANDARD_HEADERS + workbook_utils.ENGINEERING_EXTRA_HEADERS
    )
    row = [f"value-{index}" for index in range(25)]

    # Act
    workbook_utils.append_index_row(workbook, "Standards", row)

    # Assert
    assert master.max_row == 2
    assert domain.max_row == 2
    assert [cell.value for cell in master[2]] == row[:20]
    assert [cell.value for cell in domain[2]] == row


def test_get_indexed_paths_reads_master_path_column(
    workbook_utils: ModuleType,
) -> None:
    # Arrange
    workbook = openpyxl.Workbook()
    master = workbook.active
    master.title = "Master"
    master.append(workbook_utils.STANDARD_HEADERS)
    master.append(
        [
            "hash-1",
            "title",
            "author",
            "publisher",
            2024,
            None,
            "PDF",
            "Engineering/Standards/file.pdf",
            "file.pdf",
            "Engineering",
            "Geotechnical Engineering",
            "Standard",
            "Standard",
            "",
            "",
            "",
            "practitioner",
            "",
            "",
            "2026-04-26",
        ]
    )

    assert workbook_utils.get_indexed_paths(workbook) == {
        "Engineering/Standards/file.pdf"
    }


def test_domain_worksheet_constants_match_expected_structure(
    workbook_utils: ModuleType,
) -> None:
    assert workbook_utils.DOMAIN_WORKSHEETS == [
        "Ebooks",
        "Standards",
        "Magazines",
        "Misc",
    ]
    assert workbook_utils.ENHANCED_WORKSHEETS == ["Standards"]


def test_append_index_row_writes_master_and_basic_domain_rows(
    workbook_utils: ModuleType,
) -> None:
    # Arrange
    workbook = openpyxl.Workbook()
    master = workbook.active
    master.title = "Master"
    master.append(workbook_utils.STANDARD_HEADERS)
    domain = workbook.create_sheet("Ebooks")
    domain.append(workbook_utils.STANDARD_HEADERS)
    row = [f"value-{index}" for index in range(20)]

    # Act
    workbook_utils.append_index_row(workbook, "Ebooks", row)

    # Assert
    assert master.max_row == 2
    assert domain.max_row == 2
    assert [cell.value for cell in master[2]] == row
    assert [cell.value for cell in domain[2]] == row


def test_sync_notes_to_domain_worksheets_uses_path_as_join_key(
    workbook_utils: ModuleType,
) -> None:
    # Arrange
    workbook = openpyxl.Workbook()
    master = workbook.active
    master.title = "Master"
    master.append(workbook_utils.STANDARD_HEADERS)
    domain = workbook.create_sheet("Standards")
    domain.append(
        workbook_utils.STANDARD_HEADERS + workbook_utils.ENGINEERING_EXTRA_HEADERS
    )
    domain.append(
        [
            "hash",
            "title",
            "author",
            "publisher",
            2024,
            None,
            "PDF",
            "Engineering/Standards/file.pdf",
            "file.pdf",
            "Engineering",
            "Geotechnical Engineering",
            "Standard",
            "Standard",
            "",
            "",
            "",
            "practitioner",
            "",
            "NEEDS_REVIEW",
            "2026-04-26",
            "BS 1",
            "needs_review",
            "UK",
            "BSI",
            None,
        ]
    )

    # Act
    synced = workbook_utils.sync_domain_notes_by_path(
        workbook,
        {
            "Engineering/Standards/file.pdf": "DUPLICATE of Engineering/Standards/original.pdf"
        },
    )

    # Assert
    assert synced == 1
    assert (
        domain.cell(row=2, column=19).value
        == "DUPLICATE of Engineering/Standards/original.pdf"
    )


def test_summarize_workbook_reports_review_and_duplicate_counts(
    tmp_path: Path, workbook_utils: ModuleType, verify_index: ModuleType
) -> None:
    # Arrange
    index_path = tmp_path / "library-index.xlsx"
    source_folder = tmp_path / "Standards"
    source_folder.mkdir()
    (source_folder / "first.pdf").write_bytes(b"%PDF")
    (source_folder / "second.pdf").write_bytes(b"%PDF")

    workbook = openpyxl.Workbook()
    master = workbook.active
    master.title = "Master"
    master.append(workbook_utils.STANDARD_HEADERS)
    standards = workbook.create_sheet("Standards")
    standards.append(
        workbook_utils.STANDARD_HEADERS + workbook_utils.ENGINEERING_EXTRA_HEADERS
    )

    row_one = [
        "hash-1",
        "title one",
        "author",
        "publisher",
        None,
        None,
        "PDF",
        "Engineering/Standards/first.pdf",
        "first.pdf",
        "Engineering",
        "Geotechnical Engineering",
        "Standard",
        "Standard",
        "",
        "",
        "",
        "practitioner",
        "",
        "NEEDS_REVIEW: status not confirmed",
        "2026-04-26",
    ]
    row_two = [
        "hash-1",
        "title two",
        "author",
        "publisher",
        2024,
        None,
        "PDF",
        "Engineering/Standards/second.pdf",
        "second.pdf",
        "Engineering",
        "Geotechnical Engineering",
        "Standard",
        "Standard",
        "",
        "",
        "",
        "practitioner",
        "",
        "DUPLICATE of Engineering/Standards/first.pdf",
        "2026-04-26",
    ]
    master.append(row_one)
    master.append(row_two)
    standards.append([*row_one, "BS 1", "needs_review", "UK", "BSI", None])
    standards.append([*row_two, "BS 1", "needs_review", "UK", "BSI", None])
    workbook.save(index_path)

    # Act
    summary = verify_index.summarize_workbook(
        index_path=index_path,
        source_folder=source_folder,
    )

    # Assert
    assert summary["row_counts"] == {"Master": 2, "Standards": 2}
    assert summary["source_file_count"] == 2
    assert summary["needs_review_counts"] == {"Master": 1, "Standards": 1}
    assert summary["duplicate_note_counts"] == {"Master": 1, "Standards": 1}
    assert summary["missing_year_counts"] == {"Master": 1, "Standards": 1}


def test_export_needs_review_writes_review_queue_csv(
    tmp_path: Path, workbook_utils: ModuleType, export_needs_review: ModuleType
) -> None:
    # Arrange
    index_path = tmp_path / "library-index.xlsx"
    output_path = tmp_path / "needs-review.csv"
    workbook = openpyxl.Workbook()
    master = workbook.active
    master.title = "Master"
    master.append(workbook_utils.STANDARD_HEADERS)
    master.append(
        [
            "hash-1",
            "title one",
            "author",
            "publisher",
            None,
            None,
            "PDF",
            "Engineering/Standards/first.pdf",
            "first.pdf",
            "Engineering",
            "Geotechnical Engineering",
            "Standard",
            "Standard",
            "",
            "",
            "",
            "practitioner",
            "",
            "NEEDS_REVIEW: status not confirmed",
            "2026-04-26",
        ]
    )
    master.append(
        [
            "hash-2",
            "title two",
            "author",
            "publisher",
            2024,
            None,
            "PDF",
            "Engineering/Standards/second.pdf",
            "second.pdf",
            "Engineering",
            "Geotechnical Engineering",
            "Standard",
            "Standard",
            "",
            "",
            "",
            "practitioner",
            "",
            "DUPLICATE of Engineering/Standards/first.pdf",
            "2026-04-26",
        ]
    )
    workbook.save(index_path)

    # Act
    exported = export_needs_review.export_needs_review(
        index_path=index_path,
        output_path=output_path,
    )

    # Assert
    assert exported == 1
    assert "Engineering/Standards/first.pdf" in output_path.read_text(encoding="utf-8")
    assert "Engineering/Standards/second.pdf" not in output_path.read_text(
        encoding="utf-8"
    )


def test_fill_years_from_filenames_fills_missing_year_from_trailing_token(
    workbook_utils: ModuleType, enrich_safe_metadata: ModuleType
) -> None:
    # Arrange
    workbook = openpyxl.Workbook()
    master = workbook.active
    master.title = "Master"
    master.append(workbook_utils.STANDARD_HEADERS)
    master.append(
        [
            "hash-1",
            "BS EN 1997-1",
            "BSI",
            "BSI",
            None,
            None,
            "PDF",
            "Engineering/Standards/BSI/BS-EN-1997-1-2004.pdf",
            "BS-EN-1997-1-2004.pdf",
            "Engineering",
            "Geotechnical Engineering",
            "Standard",
            "Standard",
            "",
            "",
            "",
            "practitioner",
            "",
            "NEEDS_REVIEW",
            "2026-04-26",
        ]
    )

    # Act
    filled = enrich_safe_metadata.fill_years_from_filenames(workbook)

    # Assert
    assert filled == 1
    assert master.cell(row=2, column=5).value == 2004


def test_fill_years_from_filenames_skips_rows_with_existing_year(
    workbook_utils: ModuleType, enrich_safe_metadata: ModuleType
) -> None:
    # Arrange
    workbook = openpyxl.Workbook()
    master = workbook.active
    master.title = "Master"
    master.append(workbook_utils.STANDARD_HEADERS)
    master.append(
        [
            "hash-1",
            "Some Title",
            "Author",
            "Publisher",
            2020,
            None,
            "PDF",
            "Engineering/Standards/BSI/BS-EN-1997-1-2004.pdf",
            "BS-EN-1997-1-2004.pdf",
            "Engineering",
            "Geotechnical Engineering",
            "Standard",
            "Standard",
            "",
            "",
            "",
            "practitioner",
            "",
            "",
            "2026-04-26",
        ]
    )

    # Act
    filled = enrich_safe_metadata.fill_years_from_filenames(workbook)

    # Assert
    assert filled == 0
    assert master.cell(row=2, column=5).value == 2020


def test_normalize_status_values_lowercases_and_underscores(
    workbook_utils: ModuleType, enrich_safe_metadata: ModuleType
) -> None:
    # Arrange
    workbook = openpyxl.Workbook()
    master = workbook.active
    master.title = "Master"
    headers = workbook_utils.STANDARD_HEADERS + workbook_utils.ENGINEERING_EXTRA_HEADERS
    master.append(headers)
    master.append(
        [
            "hash-1",
            "BS EN 1997-1",
            "BSI",
            "BSI",
            2004,
            None,
            "PDF",
            "Engineering/Standards/BSI/BS-EN-1997-1-2004.pdf",
            "BS-EN-1997-1-2004.pdf",
            "Engineering",
            "Geotechnical Engineering",
            "Standard",
            "Standard",
            "",
            "",
            "",
            "practitioner",
            "",
            "NEEDS_REVIEW",
            "2026-04-26",
            "BS EN 1997-1:2004",
            "NEEDS_REVIEW",
            "UK",
            "BSI",
            None,
        ]
    )

    # Act
    normalized = enrich_safe_metadata.normalize_status_values(workbook)

    # Assert
    assert normalized == 1
    assert master.cell(row=2, column=22).value == "needs_review"


def test_normalize_status_values_skips_already_normalized(
    workbook_utils: ModuleType, enrich_safe_metadata: ModuleType
) -> None:
    # Arrange
    workbook = openpyxl.Workbook()
    master = workbook.active
    master.title = "Master"
    headers = workbook_utils.STANDARD_HEADERS + workbook_utils.ENGINEERING_EXTRA_HEADERS
    master.append(headers)
    master.append(
        [
            "hash-1",
            "BS EN 1997-1",
            "BSI",
            "BSI",
            2004,
            None,
            "PDF",
            "Engineering/Standards/BSI/BS-EN-1997-1-2004.pdf",
            "BS-EN-1997-1-2004.pdf",
            "Engineering",
            "Geotechnical Engineering",
            "Standard",
            "Standard",
            "",
            "",
            "",
            "practitioner",
            "",
            "NEEDS_REVIEW",
            "2026-04-26",
            "BS EN 1997-1:2004",
            "current",
            "UK",
            "BSI",
            None,
        ]
    )

    # Act
    normalized = enrich_safe_metadata.normalize_status_values(workbook)

    # Assert
    assert normalized == 0
    assert master.cell(row=2, column=22).value == "current"


def test_sync_years_to_domain_worksheets_copies_filled_years(
    workbook_utils: ModuleType, enrich_safe_metadata: ModuleType
) -> None:
    # Arrange
    workbook = openpyxl.Workbook()
    master = workbook.active
    master.title = "Master"
    master.append(workbook_utils.STANDARD_HEADERS)
    master.append(
        [
            "hash-1",
            "BS EN 1997-1",
            "BSI",
            "BSI",
            2004,
            None,
            "PDF",
            "Engineering/Standards/BSI/BS-EN-1997-1-2004.pdf",
            "BS-EN-1997-1-2004.pdf",
            "Engineering",
            "Geotechnical Engineering",
            "Standard",
            "Standard",
            "",
            "",
            "",
            "practitioner",
            "",
            "NEEDS_REVIEW",
            "2026-04-26",
        ]
    )
    domain = workbook.create_sheet("Standards")
    domain.append(
        workbook_utils.STANDARD_HEADERS + workbook_utils.ENGINEERING_EXTRA_HEADERS
    )
    domain.append(
        [
            "hash-1",
            "BS EN 1997-1",
            "BSI",
            "BSI",
            None,
            None,
            "PDF",
            "Engineering/Standards/BSI/BS-EN-1997-1-2004.pdf",
            "BS-EN-1997-1-2004.pdf",
            "Engineering",
            "Geotechnical Engineering",
            "Standard",
            "Standard",
            "",
            "",
            "",
            "practitioner",
            "",
            "NEEDS_REVIEW",
            "2026-04-26",
            "BS EN 1997-1:2004",
            "needs_review",
            "UK",
            "BSI",
            None,
        ]
    )

    # Act
    synced = enrich_safe_metadata.sync_years_to_domain_worksheets(workbook)

    # Assert
    assert synced == 1
    assert domain.cell(row=2, column=5).value == 2004


def _make_review_pack_workbook() -> openpyxl.Workbook:
    """Build a workbook with standards and non-standards rows for review-pack tests."""
    utils = load_tool_module("workbook_utils")
    workbook = openpyxl.Workbook()
    master = workbook.active
    master.title = "Master"
    master.append(utils.STANDARD_HEADERS)
    standards = workbook.create_sheet("Standards")
    standards.append(utils.STANDARD_HEADERS + utils.ENGINEERING_EXTRA_HEADERS)

    standard_row_master = [
        "hash-1",
        "BS EN 1997-1",
        "BSI",
        "BSI",
        None,
        None,
        "PDF",
        "Engineering/Standards/BSI/BS-EN-1997-1-2004.pdf",
        "BS-EN-1997-1-2004.pdf",
        "Engineering",
        "Geotechnical Engineering",
        "Standard",
        "Standard",
        "",
        "",
        "",
        "practitioner",
        "",
        "NEEDS_REVIEW: status not confirmed",
        "2026-04-26",
    ]
    non_standard_row_master = [
        "hash-2",
        "Foundation Design",
        "Author",
        "Publisher",
        2020,
        None,
        "PDF",
        "Engineering/Textbooks/foundation-design.pdf",
        "Foundation-Design_Author_2020.pdf",
        "Engineering",
        "Geotechnical Engineering",
        "Book",
        "Textbook",
        "",
        "",
        "",
        "practitioner",
        "",
        "NEEDS_REVIEW: metadata incomplete",
        "2026-04-26",
    ]
    duplicate_row_master = [
        "hash-1",
        "BS EN 1997-1 copy",
        "BSI",
        "BSI",
        2004,
        None,
        "PDF",
        "Engineering/Standards/BSI/BS-EN-1997-1-2004-copy.pdf",
        "BS-EN-1997-1-2004-copy.pdf",
        "Engineering",
        "Geotechnical Engineering",
        "Standard",
        "Standard",
        "",
        "",
        "",
        "practitioner",
        "",
        "DUPLICATE of Engineering/Standards/BSI/BS-EN-1997-1-2004.pdf",
        "2026-04-26",
    ]
    master.append(standard_row_master)
    master.append(non_standard_row_master)
    master.append(duplicate_row_master)
    standards.append(
        [*standard_row_master, "BS EN 1997-1:2004", "NEEDS_REVIEW", "UK", "BSI", None]
    )
    standards.append(
        [*duplicate_row_master, "BS EN 1997-1:2004", "NEEDS_REVIEW", "UK", "BSI", None]
    )
    return workbook


def test_export_review_pack_creates_all_csv_files(
    tmp_path: Path, export_review_pack: ModuleType
) -> None:
    # Arrange
    workbook = _make_review_pack_workbook()
    index_path = tmp_path / "library-index.xlsx"
    workbook.save(index_path)
    output_dir = tmp_path / "review-queues"

    # Act
    result = export_review_pack.export_review_pack(
        index_path=index_path,
        output_dir=output_dir,
    )

    # Assert
    assert output_dir.is_dir()
    assert (output_dir / "needs-review-all.csv").exists()
    assert (output_dir / "needs-review-standards.csv").exists()
    assert (output_dir / "needs-review-non-standards.csv").exists()
    assert (output_dir / "needs-review-missing-year.csv").exists()
    assert (output_dir / "duplicates.csv").exists()
    assert isinstance(result, dict)
    assert result["needs-review-all"] == 2
    assert result["needs-review-standards"] == 1
    assert result["needs-review-non-standards"] == 1
    assert result["needs-review-missing-year"] == 1
    assert result["duplicates"] == 1


def test_export_review_pack_standards_filter_uses_path_not_category(
    tmp_path: Path, export_review_pack: ModuleType
) -> None:
    # Arrange
    workbook = _make_review_pack_workbook()
    index_path = tmp_path / "library-index.xlsx"
    workbook.save(index_path)
    output_dir = tmp_path / "review-queues"

    # Act
    export_review_pack.export_review_pack(index_path=index_path, output_dir=output_dir)

    # Assert
    standards_csv = (output_dir / "needs-review-standards.csv").read_text(
        encoding="utf-8"
    )
    assert "Engineering/Standards/" in standards_csv
    non_standards_csv = (output_dir / "needs-review-non-standards.csv").read_text(
        encoding="utf-8"
    )
    assert "Engineering/Textbooks/" in non_standards_csv


def test_verify_index_counts_pdf_and_epub_files(
    tmp_path: Path, workbook_utils: ModuleType, verify_index: ModuleType
) -> None:
    # Arrange
    index_path = tmp_path / "library-index.xlsx"
    source_folder = tmp_path / "Library"
    source_folder.mkdir()
    (source_folder / "book.pdf").write_bytes(b"%PDF")
    (source_folder / "ebook.epub").write_bytes(b"PK")
    sub = source_folder / "Standards"
    sub.mkdir()
    (sub / "standard.pdf").write_bytes(b"%PDF")

    workbook = openpyxl.Workbook()
    master = workbook.active
    master.title = "Master"
    master.append(workbook_utils.STANDARD_HEADERS)
    workbook.save(index_path)

    # Act
    summary = verify_index.summarize_workbook(
        index_path=index_path,
        source_folder=source_folder,
    )

    # Assert
    assert summary["source_file_count"] == 3
