r"""Upsert or mark-deleted a NotebookLM notebook in an index-notebooklm.xlsx file.

The tool is path-agnostic: it operates on any index file passed via the JSON
payload's ``index_path`` key, so it can serve multiple projects and libraries.

Usage (called by Linda via run_in_terminal):

    Upsert a notebook::

        echo '{"index_path": "G:\\My Drive\\Library\\index-notebooklm.xlsx",
               "operation": "upsert",
               "notebook": {"notebook_id": "...", "title": "...", "url": "...",
                            "source_count": 3, "summary": "...", "suggested_topics": ""},
               "sources": [{"source_id": "...", "source_title": "...",
                             "source_summary": "...", "source_keywords": ""}]}' |
        python .agents/tools/library/upsert_notebooklm_index.py

    Mark a notebook as deleted::

        echo '{"index_path": "G:\\My Drive\\Library\\index-notebooklm.xlsx",
               "operation": "mark_deleted",
               "notebook_id": "abc-123"}' |
        python .agents/tools/library/upsert_notebooklm_index.py

Required JSON keys
------------------
index_path : str
    Absolute path to the target ``index-notebooklm.xlsx`` workbook.
    **Must be provided by the caller.** The tool has no built-in default path.
operation : str
    ``"upsert"`` (default if omitted), ``"mark_deleted"``, or
    ``"bulk_set_column"``.
"""

import importlib.util
import json
import sys
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

try:
    from workbook_utils import (
        WorkbookLock,
        get_header_indexes,
        save_workbook_atomically,
    )
except ModuleNotFoundError:
    util_path = Path(__file__).resolve().parent / "workbook_utils.py"
    spec = importlib.util.spec_from_file_location("workbook_utils", util_path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Could not load workbook utilities from {util_path}")
    workbook_utils = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(workbook_utils)
    WorkbookLock = workbook_utils.WorkbookLock
    get_header_indexes = workbook_utils.get_header_indexes
    save_workbook_atomically = workbook_utils.save_workbook_atomically

INDEX_PATH = Path(r"G:\My Drive\Library\index-notebooklm.xlsx")

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
DATA_ALIGN = Alignment(wrap_text=True, vertical="top")


def _apply_row_style(ws, row_idx: int) -> None:
    for cell in ws[row_idx]:
        cell.alignment = DATA_ALIGN


def upsert(
    notebook: dict,
    sources: list[dict],
    index_path: Path | None = None,
) -> str:
    """Upsert notebook and sources into the index workbook."""
    resolved_index_path = index_path or INDEX_PATH
    with WorkbookLock(resolved_index_path):
        wb = openpyxl.load_workbook(resolved_index_path)
        ws_nb = wb["notebooks"]
        ws_src = wb["sources"]

        nid = notebook["notebook_id"]
        today = date.today().isoformat()

        # --- notebooks: find existing row ---
        existing_row = None
        for row in ws_nb.iter_rows(
            min_row=2, max_row=ws_nb.max_row, min_col=1, max_col=1
        ):
            if row[0].value == nid:
                existing_row = row[0].row
                break

        nb_values = [
            nid,
            notebook["title"],
            notebook["url"],
            notebook["source_count"],
            notebook.get("summary", ""),
            notebook.get("suggested_topics", ""),
            today,
            "active",
        ]

        if existing_row:
            for col_idx, val in enumerate(nb_values, start=1):
                ws_nb.cell(row=existing_row, column=col_idx, value=val)
            _apply_row_style(ws_nb, existing_row)
            operation = "upsert (existing row overwritten)"
        else:
            ws_nb.append(nb_values)
            _apply_row_style(ws_nb, ws_nb.max_row)
            operation = "insert (new row appended)"

        # --- sources: delete existing rows for this notebook_id ---
        rows_to_delete = []
        for row in ws_src.iter_rows(
            min_row=2, max_row=ws_src.max_row, min_col=1, max_col=1
        ):
            if row[0].value == nid:
                rows_to_delete.append(row[0].row)

        for row_idx in reversed(rows_to_delete):
            ws_src.delete_rows(row_idx)

        # --- sources: insert current sources ---
        for src in sources:
            ws_src.append(
                [
                    nid,
                    notebook["title"],
                    src["source_id"],
                    src["source_title"],
                    src.get("source_summary", ""),
                    src.get("source_keywords", ""),
                ]
            )
            _apply_row_style(ws_src, ws_src.max_row)

        save_workbook_atomically(wb, resolved_index_path)

    # Verify
    wb2 = openpyxl.load_workbook(resolved_index_path)
    nb_count = wb2["notebooks"].max_row - 1
    src_count = wb2["sources"].max_row - 1
    wb2.close()

    return (
        f"{operation}\n"
        f"notebooks worksheet: {nb_count} rows (excl. header)\n"
        f"sources worksheet: {src_count} rows (excl. header)\n"
        f"deleted {len(rows_to_delete)} old source rows, inserted {len(sources)} new"
    )


def bulk_set_column(
    column_name: str,
    value: object,
    match_ids: list[str] | str,
    index_path: Path | None = None,
) -> str:
    """Set a column value on matching rows in the notebooks worksheet.

    Parameters
    ----------
    column_name:
        Header to create (if absent) or update.
    value:
        Value to write into each matched cell.
    match_ids:
        List of ``notebook_id`` strings to target, or ``"*"`` for all rows.
    index_path:
        Path to the index workbook.  Falls back to ``INDEX_PATH``.
    """
    resolved_index_path = index_path or INDEX_PATH
    with WorkbookLock(resolved_index_path):
        wb = openpyxl.load_workbook(resolved_index_path)
        ws_nb = wb["notebooks"]

        # Locate or create the target column.
        header_row = ws_nb[1]
        col_letter: int | None = None
        for cell in header_row:
            if cell.value == column_name:
                col_letter = cell.column
                break

        if col_letter is None:
            # Append new column after the last used column.
            new_col = ws_nb.max_column + 1
            header_cell = ws_nb.cell(row=1, column=new_col, value=column_name)
            header_cell.font = HEADER_FONT
            header_cell.fill = HEADER_FILL
            col_letter = new_col

        # Find the zero-based index of notebook_id column for matching.
        headers = [cell.value for cell in ws_nb[1]]
        if "notebook_id" not in headers:
            return "ERROR: 'notebook_id' column not found in notebooks worksheet"
        nid_col = headers.index("notebook_id")

        match_all = match_ids == "*"
        match_set = set() if match_all else set(match_ids)  # type: ignore[arg-type]

        updated = 0
        for row in ws_nb.iter_rows(min_row=2, max_row=ws_nb.max_row):
            row_nid = row[nid_col].value
            if match_all or row_nid in match_set:
                ws_nb.cell(row=row[0].row, column=col_letter, value=value)
                updated += 1

        save_workbook_atomically(wb, resolved_index_path)

    return (
        f"bulk_set_column complete\n"
        f"column: {column_name!r}\n"
        f"value: {value!r}\n"
        f"rows updated: {updated}"
    )


def mark_deleted(notebook_id: str, index_path: Path | None = None) -> str:
    """Mark a notebook as deleted in the index workbook."""
    resolved_index_path = index_path or INDEX_PATH
    with WorkbookLock(resolved_index_path):
        wb = openpyxl.load_workbook(resolved_index_path)
        ws_nb = wb["notebooks"]
        today = date.today().isoformat()

        header_indexes = get_header_indexes(
            ws_nb,
            ["notebook_id", "last_updated", "status"],
        )

        found = False
        for row in ws_nb.iter_rows(min_row=2, max_row=ws_nb.max_row):
            notebook_id_col = header_indexes["notebook_id"]
            if row[notebook_id_col].value == notebook_id:
                row[header_indexes["last_updated"]].value = today
                row[header_indexes["status"]].value = "deleted"
                found = True
                break

        if not found:
            return f"notebook_id {notebook_id!r} not found in {resolved_index_path}"

        save_workbook_atomically(wb, resolved_index_path)
    return f"Marked deleted: {notebook_id}"


if __name__ == "__main__":
    data = json.loads(sys.stdin.read())

    raw_path = data.get("index_path")
    if not raw_path:
        print("ERROR: 'index_path' is required in the JSON payload", file=sys.stderr)
        sys.exit(1)

    index_path = Path(raw_path)
    if not index_path.exists():
        print(f"ERROR: index file not found: {index_path}", file=sys.stderr)
        sys.exit(1)

    operation = data.get("operation", "upsert")
    if operation == "upsert":
        result = upsert(data["notebook"], data["sources"], index_path=index_path)
    elif operation == "mark_deleted":
        result = mark_deleted(data["notebook_id"], index_path=index_path)
    elif operation == "bulk_set_column":
        result = bulk_set_column(
            column_name=data["column_name"],
            value=data["value"],
            match_ids=data["match_ids"],
            index_path=index_path,
        )
    else:
        print(f"ERROR: unknown operation {operation!r}", file=sys.stderr)
        sys.exit(1)

    print(result)
