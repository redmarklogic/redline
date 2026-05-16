"""Upsert a book entry in a library-index.xlsx workbook.

The tool is path-agnostic: it operates on any workbook passed via ``index_path``
so it works across multiple projects and libraries.

Usage (called by Linda via run_in_terminal)::

    $payload = @'
    {
      "index_path": "G:\\My Drive\\Library\\library-index.xlsx",
      "library_root": "G:\\My Drive\\Library",
      "file_path": "G:\\My Drive\\Library\\Q - Science\\QA75-76 - Computer Science and Software Engineering\\Age of Invisible Machines_Wilson_2022.pdf",
      "worksheet": "Ebooks",
      "metadata": {
        "title": "Age of Invisible Machines",
        "author": "Wilson, Robb; Tyson, Josh",
        "publisher": "Wiley",
        "year": 2022,
        "edition": "",
        "domain": "Technology",
        "subdomain": "Artificial Intelligence",
        "lcc_class": "Q - Science",
        "lcc_subclass": "QA75-76 - Computer Science and Software Engineering",
        "category": "Book",
        "document_type": "Practitioner Text",
        "topics": "AI agents | orchestration | automation | conversational AI",
        "frameworks": "AI OS | multi-agent",
        "market_context": "",
        "audience": "practitioner",
        "skill_refs": "",
        "notes": ""
      }
    }
    '@
    $payload | .venv\\Scripts\\python .agents/tools/library/upsert_library_index.py

Required JSON keys
------------------
index_path : str
    Absolute path to the target ``library-index.xlsx`` workbook.
    **Must be provided by the caller.** The tool has no built-in default path.
library_root : str, optional
    Absolute path to the library root folder (used to derive the relative ``path``
    column value). Defaults to the parent directory of ``index_path``.
file_path : str
    Absolute path to the PDF or EPUB file on disk. Used to compute SHA-256
    and derive ``path``, ``canonical_filename``, and ``format``.
worksheet : str
    Target domain worksheet name. One of: ``Ebooks``, ``Standards``,
    ``Magazines``, ``Misc``.
metadata : dict
    All metadata fields. The caller supplies every field; the tool derives
    ``sha256``, ``path``, ``canonical_filename``, ``format``, and
    ``last_updated`` automatically from ``file_path``.

Idempotency
-----------
The relative ``path`` is the resume/idempotency key (matching the convention in
``workbook_utils``). Calling the tool twice for the same file overwrites the
existing row rather than appending a duplicate.
"""

from __future__ import annotations

import hashlib
import importlib.util
import json
import sys
from datetime import date
from pathlib import Path

import openpyxl

try:
    from workbook_utils import (
        STANDARD_HEADERS,
        WorkbookLock,
        get_header_indexes,
        save_workbook_atomically,
    )
except ModuleNotFoundError:
    util_path = Path(__file__).resolve().parent / "workbook_utils.py"
    spec = importlib.util.spec_from_file_location("workbook_utils", util_path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Could not load workbook utilities from {util_path}")
    workbook_utils = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(workbook_utils)
    STANDARD_HEADERS = workbook_utils.STANDARD_HEADERS
    WorkbookLock = workbook_utils.WorkbookLock
    get_header_indexes = workbook_utils.get_header_indexes
    save_workbook_atomically = workbook_utils.save_workbook_atomically


def _sha256(file_path: Path) -> str:
    h = hashlib.sha256()
    with file_path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(65536), b""):
            h.update(chunk)
    return h.hexdigest().upper()


def _build_row(
    file_path: Path,
    library_root: Path,
    metadata: dict,
) -> list[object]:
    """Assemble a full STANDARD_HEADERS row from file_path + caller metadata."""
    rel_path = str(file_path.relative_to(library_root))
    fmt = "EPUB" if file_path.suffix.lower() == ".epub" else "PDF"
    today = date.today().isoformat()

    return [
        _sha256(file_path),
        metadata.get("title", ""),
        metadata.get("author", ""),
        metadata.get("publisher", ""),
        metadata.get("year") or None,
        metadata.get("edition") or None,
        fmt,
        rel_path,
        file_path.name,
        metadata.get("domain", ""),
        metadata.get("subdomain", ""),
        metadata.get("lcc_class", ""),
        metadata.get("lcc_subclass", ""),
        metadata.get("category", ""),
        metadata.get("document_type", ""),
        metadata.get("topics", ""),
        metadata.get("frameworks", ""),
        metadata.get("market_context", ""),
        metadata.get("audience", ""),
        metadata.get("skill_refs", ""),
        metadata.get("notes", ""),
        today,
    ]


def upsert(
    index_path: Path,
    library_root: Path,
    file_path: Path,
    worksheet_name: str,
    metadata: dict,
) -> str:
    """Upsert one book row into Master and the target domain worksheet."""
    rel_path = str(file_path.relative_to(library_root))
    row = _build_row(file_path, library_root, metadata)

    with WorkbookLock(index_path):
        wb = openpyxl.load_workbook(index_path)
        operation = _upsert_row(wb, "Master", rel_path, row[: len(STANDARD_HEADERS)])
        _upsert_row(wb, worksheet_name, rel_path, row)
        save_workbook_atomically(wb, index_path)

    master_count = (
        openpyxl.load_workbook(index_path, read_only=True)["Master"].max_row - 1
    )
    return (
        f"{operation}\n"
        f"file: {file_path.name}\n"
        f"worksheet: {worksheet_name}\n"
        f"Master rows (excl. header): {master_count}"
    )


def _upsert_row(
    wb: openpyxl.Workbook,
    worksheet_name: str,
    rel_path: str,
    row: list[object],
) -> str:
    ws = wb[worksheet_name]
    header_indexes = get_header_indexes(ws, ["path"])
    path_col = header_indexes["path"] + 1  # 1-based

    for existing_row in ws.iter_rows(min_row=2):
        if existing_row[path_col - 1].value == rel_path:
            for col_idx, val in enumerate(row, start=1):
                ws.cell(row=existing_row[0].row, column=col_idx, value=val)
            return "upsert (existing row overwritten)"

    ws.append(row)
    return "insert (new row appended)"


if __name__ == "__main__":
    data = json.loads(sys.stdin.read())

    raw_index = data.get("index_path")
    if not raw_index:
        print("ERROR: 'index_path' is required in the JSON payload", file=sys.stderr)
        sys.exit(1)

    index_path = Path(raw_index)
    if not index_path.exists():
        print(f"ERROR: index file not found: {index_path}", file=sys.stderr)
        sys.exit(1)

    library_root = Path(data.get("library_root") or index_path.parent)

    raw_file = data.get("file_path")
    if not raw_file:
        print("ERROR: 'file_path' is required in the JSON payload", file=sys.stderr)
        sys.exit(1)

    file_path = Path(raw_file)
    if not file_path.exists():
        print(f"ERROR: file not found: {file_path}", file=sys.stderr)
        sys.exit(1)

    worksheet = data.get("worksheet", "Ebooks")
    metadata = data.get("metadata", {})

    result = upsert(index_path, library_root, file_path, worksheet, metadata)
    print(result)
