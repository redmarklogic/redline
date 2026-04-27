r"""One-off script: create the Standards worksheet in library-index.xlsx.

Run once from the repo root:
    .venv\Scripts\python.exe .agents\tools\library\create_standards_worksheet.py
"""

import pathlib
import sys

import openpyxl

sys.path.insert(0, str(pathlib.Path(__file__).parent))
from workbook_utils import (
    ENGINEERING_EXTRA_HEADERS,
    INDEX_PATH,
    STANDARD_HEADERS,
    WorkbookLock,
    save_workbook_atomically,
)


def create_standards_worksheet() -> None:
    """Create the Standards worksheet if it is missing."""
    with WorkbookLock(INDEX_PATH):
        workbook = openpyxl.load_workbook(INDEX_PATH)
        if "Standards" in workbook.sheetnames:
            print("'Standards' worksheet already exists — no changes made.")
            return
        worksheet = workbook.create_sheet("Standards")
        headers = STANDARD_HEADERS + ENGINEERING_EXTRA_HEADERS
        worksheet.append(headers)
        save_workbook_atomically(workbook, INDEX_PATH)

    print(f"Created 'Standards' worksheet with {len(headers)} columns:")
    print("  " + ", ".join(headers))


if __name__ == "__main__":
    create_standards_worksheet()
