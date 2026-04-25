"""Create a fresh library-index.xlsx with all required worksheets and headers.

Run from the repo root:
    .venv\Scripts\python.exe .agents\tools\library\create_fresh_index.py

Creates G:\My Drive\Library\library-index.xlsx with:
  - Master worksheet (standard 20 columns)
  - Engineering worksheet (20 standard + 5 engineering-specific columns)
  - Standards worksheet (20 standard + 5 engineering-specific columns)

WARNING: This overwrites any existing library-index.xlsx.
"""

import pathlib

import openpyxl
from openpyxl.styles import Font, PatternFill

INDEX_PATH = pathlib.Path(r"G:\My Drive\Library\library-index.xlsx")

STANDARD_HEADERS = [
    "sha256",
    "title",
    "author",
    "publisher",
    "year",
    "edition",
    "format",
    "path",
    "canonical_filename",
    "domain",
    "subdomain",
    "category",
    "document_type",
    "topics",
    "frameworks",
    "market_context",
    "audience",
    "skill_refs",
    "notes",
    "last_updated",
]

ENGINEERING_EXTRA_HEADERS = [
    "standard_code",
    "status",
    "jurisdiction",
    "issuing_body",
    "superseded_by",
]

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _write_header_row(
    ws: openpyxl.worksheet.worksheet.Worksheet, headers: list[str]
) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws.freeze_panes = "A2"


wb = openpyxl.Workbook()

# Master worksheet
ws_master = wb.active
ws_master.title = "Master"
_write_header_row(ws_master, STANDARD_HEADERS)

# Engineering worksheet (standard + engineering extra)
ws_eng = wb.create_sheet("Engineering")
_write_header_row(ws_eng, STANDARD_HEADERS + ENGINEERING_EXTRA_HEADERS)

# Standards worksheet (standard + engineering extra)
ws_standards = wb.create_sheet("Standards")
_write_header_row(ws_standards, STANDARD_HEADERS + ENGINEERING_EXTRA_HEADERS)

wb.save(INDEX_PATH)
print(f"Created: {INDEX_PATH}")
print(f"Worksheets: {wb.sheetnames}")
print("Headers written. Ready for Phase 1 (batch_index.py).")
