r"""Ebooks worksheet audit: filename convention, created_date, subfolder review, provisional flags.

Permanent tool — do NOT run `tmp_*.py` equivalents.

Run:
    cd <repo-root>
    .\.venv\Scripts\python .agents\tools\library\audit_ebooks.py

Four tasks:
  1. Canonical filename audit  — flags rows with NEEDS_REVIEW: filename
  2. created_date backfill     — counts missing values; flags rows
  3. Per-subfolder review      — split candidates / informational / healthy
  4. Provisional listing       — entries blocked from NotebookLM upload

Updates notes columns in Ebooks and Master under WorkbookLock.
Does NOT rename or move physical files. Does NOT upload to NotebookLM.
"""

import pathlib
import re
import sys
from collections import Counter
from datetime import date

import openpyxl

sys.path.insert(0, str(pathlib.Path(__file__).parent))
from workbook_utils import INDEX_PATH, WorkbookLock, save_workbook_atomically

EBOOK_WORKSHEET = "Ebooks"
TODAY = date.today()
VELOCITY_WINDOW_DAYS = 90


# ---------------------------------------------------------------------------
# Filename validation
# ---------------------------------------------------------------------------


def is_canonical_filename(filename: str) -> bool:  # noqa: PLR0911
    """Return True if filename matches Full-Title_AuthorSurname_YYYY.(pdf|epub).

    Compliant filename requirements:
    - Extension is .pdf or .epub (case-insensitive, but lower-case is canonical)
    - No spaces in the base name (spaces indicate wrong separator)
    - At least 3 underscore-separated tokens: title part, author, year
    - Last token is exactly 4 digits (year)
    - Second-to-last token is a single word representing the author surname
      (letters, hyphens, apostrophes, periods — e.g. van-der-Berg, O'Brien)
    - Base name is not entirely upper-case (all-caps is non-canonical)
    """
    if not filename:
        return False
    m = re.match(r"^(.+)\.(pdf|epub)$", filename, re.IGNORECASE)
    if not m:
        return False
    name = m.group(1)
    # Spaces → wrong separator
    if " " in name:
        return False
    # All-caps base name
    if name == name.upper() and any(c.isalpha() for c in name):
        return False
    tokens = name.split("_")
    if len(tokens) < 3:
        return False
    # Year: exactly 4 digits
    if not re.match(r"^\d{4}$", tokens[-1]):
        return False
    # Author surname: letters + optional hyphens, apostrophes, periods
    if not re.match(r"^[A-Za-z][A-Za-z\-'.]*$", tokens[-2]):
        return False
    # Title part: at least one non-empty token
    title_part = "_".join(tokens[:-2])
    return bool(title_part.strip())


# ---------------------------------------------------------------------------
# Main audit
# ---------------------------------------------------------------------------


def audit_ebooks(index_path: pathlib.Path = INDEX_PATH) -> None:  # noqa: PLR0912, PLR0915
    """Run full Ebooks audit; update notes columns; print structured report."""
    # --- Pre-update verification ---
    pre_wb = openpyxl.load_workbook(index_path, read_only=True, data_only=True)
    pre_ebooks_rows = max(pre_wb[EBOOK_WORKSHEET].max_row - 1, 0)
    pre_master_rows = max(pre_wb["Master"].max_row - 1, 0)
    pre_needs_review = _count_column_value(
        pre_wb[EBOOK_WORKSHEET], "notes", "NEEDS_REVIEW"
    )
    pre_wb.close()

    # --- Load for writing ---
    workbook = openpyxl.load_workbook(index_path)
    ws_ebooks = workbook[EBOOK_WORKSHEET]
    ws_master = workbook["Master"]

    # Resolve Ebooks column indexes
    ebook_headers = [cell.value for cell in ws_ebooks[1]]

    def _col(name: str) -> int | None:
        return ebook_headers.index(name) if name in ebook_headers else None

    notes_col = _col("notes")
    filename_col = _col("canonical_filename")
    subdomain_col = _col("subdomain")
    topics_col = _col("topics")
    title_col = _col("title")
    path_col = _col("path")
    created_date_col = _col("created_date")
    has_created_date = created_date_col is not None

    if notes_col is None or filename_col is None:
        print(
            "ERROR: Required columns (notes, canonical_filename) missing from Ebooks worksheet."
        )
        return

    # --- Accumulators ---
    filename_compliant = 0
    filename_flagged = 0
    already_flagged_filename = 0
    created_date_missing = 0
    already_flagged_cd = 0
    rows_updated = 0

    subdomain_counts: Counter[str] = Counter()
    subdomain_topics: dict[str, Counter[str]] = {}
    subdomain_provisional: Counter[str] = Counter()
    subdomain_velocity: Counter[str] = Counter()  # books added in last 90 days
    provisional_rows: list[dict[str, str]] = []

    # Collect path→new_notes for Master sync
    path_to_new_notes: dict[str, str] = {}

    # --- Row loop ---
    for row in ws_ebooks.iter_rows(min_row=2):
        canonical_fn = str(row[filename_col].value or "").strip()
        original_notes = str(row[notes_col].value or "").strip()
        subdomain_val = (
            str(row[subdomain_col].value or "").strip()
            if subdomain_col is not None
            else ""
        )
        topics_val = (
            str(row[topics_col].value or "").strip() if topics_col is not None else ""
        )
        title_val = (
            str(row[title_col].value or "").strip() if title_col is not None else ""
        )
        path_val = (
            str(row[path_col].value or "").strip() if path_col is not None else ""
        )

        # Subdomain stats (computed on original notes state)
        if subdomain_val:
            subdomain_counts[subdomain_val] += 1
            if subdomain_val not in subdomain_topics:
                subdomain_topics[subdomain_val] = Counter()
            for tag in [t.strip() for t in topics_val.split(";") if t.strip()]:
                subdomain_topics[subdomain_val][tag] += 1

        # Provisional tracking
        if "PROVISIONAL_CLASSIFICATION" in original_notes:
            subdomain_provisional[subdomain_val] += 1
            provisional_rows.append(
                {
                    "title": title_val,
                    "subdomain": subdomain_val,
                    "topics": topics_val,
                    "notes": original_notes,
                }
            )

        # Velocity tracking (requires created_date column)
        if has_created_date:
            cd_raw = row[created_date_col].value
            if cd_raw:
                try:
                    cd_date = date.fromisoformat(str(cd_raw)[:10])
                    if (TODAY - cd_date).days <= VELOCITY_WINDOW_DAYS:
                        subdomain_velocity[subdomain_val] += 1
                except ValueError:
                    pass

        # --- Compute flags (based on ORIGINAL notes state) ---
        new_flags: list[str] = []

        # Task 1: canonical filename
        if not is_canonical_filename(canonical_fn):
            filename_flagged += 1
            flag1 = "NEEDS_REVIEW: filename"
            if flag1 not in original_notes:
                new_flags.append(flag1)
            else:
                already_flagged_filename += 1
        else:
            filename_compliant += 1

        # Task 2: created_date
        if has_created_date:
            cd_val = row[created_date_col].value
            if not cd_val:
                created_date_missing += 1
                if "created_date missing" not in original_notes:
                    # Format per spec: empty notes → full flag; notes present → append suffix
                    if original_notes:
                        new_flags.append("created_date missing")
                    else:
                        new_flags.append("NEEDS_REVIEW: created_date missing")
                else:
                    already_flagged_cd += 1

        # --- Apply flags ---
        if new_flags:
            additions = "; ".join(new_flags)
            new_notes = (
                (original_notes + "; " + additions) if original_notes else additions
            )
            row[notes_col].value = new_notes
            rows_updated += 1
            if path_val:
                path_to_new_notes[path_val] = new_notes

    # --- Sync updated notes to Master worksheet ---
    master_headers = [cell.value for cell in ws_master[1]]
    master_path_col = master_headers.index("path") if "path" in master_headers else None
    master_notes_col = (
        master_headers.index("notes") if "notes" in master_headers else None
    )
    master_rows_synced = 0

    if (
        master_path_col is not None
        and master_notes_col is not None
        and path_to_new_notes
    ):
        for row in ws_master.iter_rows(min_row=2):
            p = str(row[master_path_col].value or "").strip()
            if p in path_to_new_notes:
                row[master_notes_col].value = path_to_new_notes[p]
                master_rows_synced += 1

    # --- Save atomically ---
    with WorkbookLock(index_path):
        save_workbook_atomically(workbook, index_path)
    workbook.close()

    # --- Post-update verification ---
    post_wb = openpyxl.load_workbook(index_path, read_only=True, data_only=True)
    post_ebooks_rows = max(post_wb[EBOOK_WORKSHEET].max_row - 1, 0)
    post_master_rows = max(post_wb["Master"].max_row - 1, 0)
    post_needs_review = _count_column_value(
        post_wb[EBOOK_WORKSHEET], "notes", "NEEDS_REVIEW"
    )
    post_wb.close()

    # --- Print structured report ---
    _print_report(
        pre_ebooks_rows=pre_ebooks_rows,
        pre_master_rows=pre_master_rows,
        post_ebooks_rows=post_ebooks_rows,
        post_master_rows=post_master_rows,
        pre_needs_review=pre_needs_review,
        post_needs_review=post_needs_review,
        rows_updated=rows_updated,
        master_rows_synced=master_rows_synced,
        filename_compliant=filename_compliant,
        filename_flagged=filename_flagged,
        already_flagged_filename=already_flagged_filename,
        has_created_date=has_created_date,
        created_date_missing=created_date_missing,
        already_flagged_cd=already_flagged_cd,
        subdomain_counts=subdomain_counts,
        subdomain_topics=subdomain_topics,
        subdomain_provisional=subdomain_provisional,
        subdomain_velocity=subdomain_velocity,
        provisional_rows=provisional_rows,
    )


# ---------------------------------------------------------------------------
# Report printer
# ---------------------------------------------------------------------------


def _print_report(  # noqa: PLR0913, PLR0912, PLR0915
    *,
    pre_ebooks_rows: int,
    pre_master_rows: int,
    post_ebooks_rows: int,
    post_master_rows: int,
    pre_needs_review: int,
    post_needs_review: int,
    rows_updated: int,
    master_rows_synced: int,
    filename_compliant: int,
    filename_flagged: int,
    already_flagged_filename: int,
    has_created_date: bool,
    created_date_missing: int,
    already_flagged_cd: int,
    subdomain_counts: Counter[str],
    subdomain_topics: dict[str, Counter[str]],
    subdomain_provisional: Counter[str],
    subdomain_velocity: Counter[str],
    provisional_rows: list[dict[str, str]],
) -> None:
    total_rows = filename_compliant + filename_flagged
    sep = "=" * 72

    print(sep)
    print("  EBOOKS LIBRARY AUDIT REPORT")
    print(f"  Generated: {TODAY.isoformat()}")
    print(sep)

    # --- Workbook verification ---
    print("\nWORKBOOK VERIFICATION")
    print("-" * 40)
    print(
        f"  Ebooks rows   pre / post : {pre_ebooks_rows} / {post_ebooks_rows}  (delta {post_ebooks_rows - pre_ebooks_rows:+d})"
    )
    print(
        f"  Master rows   pre / post : {pre_master_rows} / {post_master_rows}  (delta {post_master_rows - pre_master_rows:+d})"
    )
    print(
        f"  NEEDS_REVIEW  pre / post : {pre_needs_review} / {post_needs_review}  (delta {post_needs_review - pre_needs_review:+d})"
    )
    print(f"  Ebooks notes rows updated: {rows_updated}")
    print(f"  Master notes rows synced : {master_rows_synced}")

    # --- Task 1 ---
    print("\n" + sep)
    print("  TASK 1 — CANONICAL FILENAME AUDIT")
    print(sep)
    pct_flagged = (filename_flagged / total_rows * 100) if total_rows else 0.0
    print(f"  Total Ebook rows examined  : {total_rows}")
    print(f"  Compliant                  : {filename_compliant}")
    print(f"  Non-compliant (flagged)    : {filename_flagged}  ({pct_flagged:.1f}%)")
    print(
        f"    of which already flagged : {already_flagged_filename}  (pre-existing NEEDS_REVIEW: filename)"
    )
    print(
        f"    newly flagged this run   : {filename_flagged - already_flagged_filename}"
    )

    # --- Task 2 ---
    print("\n" + sep)
    print("  TASK 2 — created_date BACKFILL STATUS")
    print(sep)
    if has_created_date:
        print("  created_date column present : YES")
        print(f"  Rows with empty created_date: {created_date_missing} of {total_rows}")
        print(f"    of which already flagged  : {already_flagged_cd}")
        print(
            f"    newly flagged this run    : {created_date_missing - already_flagged_cd}"
        )
        if created_date_missing > 0:
            print("\n  ACTION REQUIRED: Founder decision needed on backfill strategy")
            print(f"  for {created_date_missing} rows. Suggested options:")
            print(
                "    A) Leave blank — created_date is acquisition date, not publication date"
            )
            print(
                f"    B) Use today ({TODAY.isoformat()}) as a floor date for all existing rows"
            )
            print(
                "    C) Attempt to infer from git history / file timestamps (manual effort)"
            )
    else:
        print("  created_date column present : NO")
        print(f"  Estimated missing           : all {total_rows} rows")
        print("\n  ACTION REQUIRED:")
        print("    1. Add 'created_date' column header to Ebooks and Master worksheets")
        print("    2. Re-run this script after adding the column header")
        print(
            f"    3. Founder to decide backfill strategy for all {total_rows} existing rows"
        )

    # --- Task 3 ---
    print("\n" + sep)
    print("  TASK 3 — PER-SUBFOLDER REVIEW")
    print(sep)
    print(f"  Distinct subdomains: {len(subdomain_counts)}")

    split_candidates: list[dict] = []
    informational: list[dict] = []
    healthy: list[dict] = []

    for subdomain, count in sorted(subdomain_counts.items(), key=lambda x: -x[1]):
        provisional_count = subdomain_provisional.get(subdomain, 0)
        velocity_count = subdomain_velocity.get(subdomain, 0)
        top_topics = subdomain_topics.get(subdomain, Counter()).most_common(5)
        entry = {
            "subdomain": subdomain,
            "count": count,
            "provisional": provisional_count,
            "velocity": velocity_count,
            "top_topics": top_topics,
        }
        if count >= 15 or provisional_count > 0:
            split_candidates.append(entry)
        elif count >= 10:
            informational.append(entry)
        else:
            healthy.append(entry)

    if split_candidates:
        print("\n  SPLIT CANDIDATES  (count >= 15 OR provisional books present)")
        print("  ----------------------------------------------------------------")
        for e in split_candidates:
            reasons = []
            if e["count"] >= 15:
                reasons.append(f"count={e['count']}")
            if e["provisional"] > 0:
                reasons.append(f"provisional={e['provisional']}")
            if e["velocity"] >= 5:
                reasons.append(f"velocity={e['velocity']} in 90d")
            print(f"\n  [{', '.join(reasons)}]")
            print(f"  Subfolder review: {e['subdomain']}")
            print(f"  - Total books:                  {e['count']}")
            print(f"  - Books added in last 90 days:  {e['velocity']}")
            topics_str = (
                ", ".join(f"{t}({n})" for t, n in e["top_topics"])
                if e["top_topics"]
                else "(none)"
            )
            print(f"  - Top topics:                   {topics_str}")
            print(f"  - Provisional classification:   {e['provisional']}")
            print(f"  - Recommendation trigger:       {', '.join(reasons)}")

    if informational:
        print("\n  INFORMATIONAL  (count 10-14, no provisionals)")
        print("  ------------------------------------------------")
        for e in informational:
            topics_str = (
                ", ".join(f"{t}({n})" for t, n in e["top_topics"])
                if e["top_topics"]
                else "(none)"
            )
            print(f"\n  Subfolder review: {e['subdomain']}")
            print(f"  - Total books: {e['count']}  |  Top topics: {topics_str}")

    if healthy:
        print("\n  HEALTHY  (count < 10, no provisionals)")
        print("  ---------------------------------------")
        for e in healthy:
            print(f"    [{e['count']:2d}] {e['subdomain']}")

    # --- Task 4 ---
    print("\n" + sep)
    print("  TASK 4 — PROVISIONAL_CLASSIFICATION ENTRIES")
    print(sep)
    if provisional_rows:
        print(
            f"  {len(provisional_rows)} row(s) carry PROVISIONAL_CLASSIFICATION flag."
        )
        print(
            "  These MUST NOT be uploaded to NotebookLM until cleared by domain agent."
        )
        print()
        for i, r in enumerate(provisional_rows, 1):
            print(f"  [{i}]")
            print(f"    Title    : {r['title'] or '(blank)'}")
            print(f"    Subdomain: {r['subdomain'] or '(blank)'}")
            print(f"    Topics   : {r['topics'] or '(blank)'}")
            print(f"    Notes    : {r['notes']}")
            print()
    else:
        print("  None found. No entries blocking NotebookLM uploads.")

    print(sep)
    print("  AUDIT COMPLETE — no physical files were renamed or moved.")
    print("  No uploads to NotebookLM were performed.")
    print(sep)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _count_column_value(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    col_name: str,
    token: str,
) -> int:
    headers = [cell.value for cell in worksheet[1]]
    if col_name not in headers:
        return 0
    idx = headers.index(col_name)
    return sum(
        1
        for row in worksheet.iter_rows(min_row=2)
        if row[idx].value and token in str(row[idx].value)
    )


if __name__ == "__main__":
    audit_ebooks()
