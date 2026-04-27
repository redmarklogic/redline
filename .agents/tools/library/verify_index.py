"""Verify and summarize the digital-library workbook."""

import pathlib
import sys

import openpyxl

sys.path.insert(0, str(pathlib.Path(__file__).parent))
from workbook_utils import INDEX_PATH, LIBRARY_ROOT, get_header_indexes


def summarize_workbook(
    *,
    index_path: pathlib.Path = INDEX_PATH,
    source_folder: pathlib.Path | None = None,
) -> dict[str, object]:
    """Return row, review, duplicate, and missing-year counts for a workbook."""
    workbook = openpyxl.load_workbook(index_path, read_only=True, data_only=True)
    try:
        worksheets = [worksheet.title for worksheet in workbook.worksheets]
        row_counts = {
            worksheet.title: _count_data_rows(worksheet)
            for worksheet in workbook.worksheets
        }
        needs_review_counts = {
            worksheet.title: _count_rows_containing(worksheet, "notes", "NEEDS_REVIEW")
            for worksheet in workbook.worksheets
            if _has_header(worksheet, "notes")
        }
        duplicate_note_counts = {
            worksheet.title: _count_rows_containing(worksheet, "notes", "DUPLICATE of")
            for worksheet in workbook.worksheets
            if _has_header(worksheet, "notes")
        }
        missing_year_counts = {
            worksheet.title: _count_missing_values(worksheet, "year")
            for worksheet in workbook.worksheets
            if _has_header(worksheet, "year")
        }
    finally:
        workbook.close()

    return {
        "worksheets": worksheets,
        "row_counts": row_counts,
        "source_file_count": _count_source_files(source_folder),
        "needs_review_counts": needs_review_counts,
        "duplicate_note_counts": duplicate_note_counts,
        "missing_year_counts": missing_year_counts,
    }


def print_summary(summary: dict[str, object]) -> None:
    """Print a stable human-readable workbook verification summary."""
    print(f"Worksheets: {summary['worksheets']}")
    print(f"Row counts: {summary['row_counts']}")
    if summary["source_file_count"] is not None:
        print(f"Source file count (PDF+EPUB): {summary['source_file_count']}")
    print(f"NEEDS_REVIEW counts: {summary['needs_review_counts']}")
    print(f"Duplicate note counts: {summary['duplicate_note_counts']}")
    print(f"Missing year counts: {summary['missing_year_counts']}")


def _count_data_rows(worksheet: openpyxl.worksheet.worksheet.Worksheet) -> int:
    return max(worksheet.max_row - 1, 0)


def _count_source_files(source_folder: pathlib.Path | None) -> int | None:
    if source_folder is None:
        return None
    extensions = {".pdf", ".epub"}
    return sum(
        1
        for file_path in source_folder.rglob("*")
        if file_path.suffix.lower() in extensions and file_path.name != "desktop.ini"
    )


def _count_rows_containing(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    header_name: str,
    token: str,
) -> int:
    header_indexes = get_header_indexes(worksheet, [header_name])
    column_idx = header_indexes[header_name]
    return sum(
        1
        for row in worksheet.iter_rows(min_row=2)
        if row[column_idx].value and token in str(row[column_idx].value)
    )


def _count_missing_values(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    header_name: str,
) -> int:
    header_indexes = get_header_indexes(worksheet, [header_name])
    column_idx = header_indexes[header_name]
    return sum(
        1
        for row in worksheet.iter_rows(min_row=2)
        if row[column_idx].value in (None, "")
    )


def _has_header(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    header_name: str,
) -> bool:
    headers = [cell.value for cell in worksheet[1]]
    return header_name in headers


if __name__ == "__main__":
    print_summary(summarize_workbook(index_path=INDEX_PATH, source_folder=LIBRARY_ROOT))
