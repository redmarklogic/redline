"""Safe mechanical enrichment for the digital-library workbook.

Operations here are domain-judgment-free: they derive values from filenames,
normalize vocabulary, and sync fields across worksheets.
"""

import pathlib
import re
import sys

import openpyxl

sys.path.insert(0, str(pathlib.Path(__file__).parent))
from workbook_utils import (
    INDEX_PATH,
    WorkbookLock,
    get_header_indexes,
    save_workbook_atomically,
)

YEAR_PATTERN = re.compile(r"(?:^|[_\-])(\d{4})(?:[_\-.]|$)")
VALID_STATUSES = {"current", "superseded", "withdrawn", "draft", "needs_review"}


def fill_years_from_filenames(workbook: openpyxl.Workbook) -> int:
    """Fill missing year values in Master from unambiguous trailing filename tokens."""
    master = workbook["Master"]
    header_indexes = get_header_indexes(master, ["year", "canonical_filename"])
    year_idx = header_indexes["year"]
    filename_idx = header_indexes["canonical_filename"]
    filled = 0
    for row in master.iter_rows(min_row=2):
        if row[year_idx].value not in (None, ""):
            continue
        filename = str(row[filename_idx].value or "")
        matches = YEAR_PATTERN.findall(filename)
        if matches:
            row[year_idx].value = int(matches[-1])
            filled += 1
    return filled


def normalize_status_values(workbook: openpyxl.Workbook) -> int:
    """Normalize status column values across all worksheets that have a status header."""
    normalized = 0
    for worksheet in workbook.worksheets:
        headers = [cell.value for cell in worksheet[1]]
        if "status" not in headers:
            continue
        status_idx = headers.index("status")
        for row in worksheet.iter_rows(min_row=2):
            raw = row[status_idx].value
            if raw is None or raw == "":
                continue
            canonical = str(raw).strip().lower().replace(" ", "_")
            if canonical != str(raw):
                row[status_idx].value = canonical
                normalized += 1
    return normalized


def sync_years_to_domain_worksheets(workbook: openpyxl.Workbook) -> int:
    """Copy year values from Master rows to matching domain worksheet rows by path."""
    master = workbook["Master"]
    master_indexes = get_header_indexes(master, ["path", "year"])
    path_idx = master_indexes["path"]
    year_idx = master_indexes["year"]
    years_by_path: dict[str, int] = {}
    for row in master.iter_rows(min_row=2):
        path_val = str(row[path_idx].value) if row[path_idx].value else ""
        year_val = row[year_idx].value
        if path_val and year_val not in (None, ""):
            years_by_path[path_val] = int(year_val)

    synced = 0
    for worksheet in workbook.worksheets:
        if worksheet.title == "Master":
            continue
        headers = [cell.value for cell in worksheet[1]]
        if "path" not in headers or "year" not in headers:
            continue
        ws_path_idx = headers.index("path")
        ws_year_idx = headers.index("year")
        for row in worksheet.iter_rows(min_row=2):
            row_path = str(row[ws_path_idx].value) if row[ws_path_idx].value else ""
            master_year = years_by_path.get(row_path)
            if master_year is not None and row[ws_year_idx].value in (None, ""):
                row[ws_year_idx].value = master_year
                synced += 1
    return synced


if __name__ == "__main__":
    with WorkbookLock(INDEX_PATH):
        wb = openpyxl.load_workbook(INDEX_PATH)
        years_filled = fill_years_from_filenames(wb)
        statuses_normalized = normalize_status_values(wb)
        years_synced = sync_years_to_domain_worksheets(wb)
        save_workbook_atomically(wb, INDEX_PATH)
    print(f"Years filled: {years_filled}")
    print(f"Statuses normalized: {statuses_normalized}")
    print(f"Years synced to domain worksheets: {years_synced}")
